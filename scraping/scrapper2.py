from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import openpyxl
from selenium.webdriver.common.by import By

path = r'C:\Users\virin\OneDrive\Desktop\Files\practice\Book 1.xlsx'
wb = openpyxl.load_workbook(path)
sh1 = wb.active

url = "https://books.toscrape.com/"

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
chrome_options.binary_location = r'Z:\Rgit(P)\chrome-win64\chrome.exe'
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install(),options=chrome_options))
driver.get(url)
driver.maximize_window()
#-----------------------------------------------------------------------------------------------------------------------------------------------
page_title = driver.find_element(By.XPATH, '//li[@class="current"]').text
print(page_title)
page_last_num = page_title.split()[-1]
page_count = int(page_last_num)
print(page_count)

for p in range(1,page_count+1):
    if p==1:
        books = driver.find_elements(By.XPATH,'//article//a[@title]')
        for book in books:
            book.click()
            book_title = driver.find_element(By.XPATH, '//h1').text
            print(book_title)
            table_rows = len(driver.find_elements(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr'))
            print(table_rows)
            UPC = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[1]/td').text
            product_type = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[2]/td').text
            price_excltax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[3]/td').text
            price_incltax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[4]/td').text
            tax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[5]/td').text
            availability = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[6]/td').text
            number_of_reviews = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[7]/td').text
            excel_max_row = sh1.max_row
            sh1.cell(row =excel_max_row+1, column =1 ).value = book_title
            sh1.cell(row =excel_max_row+1, column =2 ).value = UPC
            sh1.cell(row =excel_max_row+1, column =3 ).value = product_type
            sh1.cell(row =excel_max_row+1, column =4 ).value = price_excltax
            sh1.cell(row =excel_max_row+1, column =5 ).value = price_incltax
            sh1.cell(row =excel_max_row+1, column =6 ).value = tax
            sh1.cell(row =excel_max_row+1, column =7 ).value = availability
            sh1.cell(row =excel_max_row+1, column =8 ).value = number_of_reviews
            wb.save(path)
            driver.back()
    else:
        driver.find_element(By.XPATH,"//*[contains(@href, 'page-"+str(p)+"')]").click()
        books = driver.find_elements(By.XPATH,'//article//a[@title]')
        for book in books:
            book.click()
            book_title = driver.find_element(By.XPATH, '//h1').text
            print(book_title)
            table_rows = len(driver.find_elements(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr'))
            print(table_rows)
            UPC = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[1]/td').text
            product_type = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[2]/td').text
            price_excltax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[3]/td').text
            price_incltax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[4]/td').text
            tax = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[5]/td').text
            availability = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[6]/td').text
            number_of_reviews = driver.find_element(By.XPATH,'//*[@class = "table table-striped"]/tbody/tr[7]/td').text
            excel_max_row = sh1.max_row
            sh1.cell(row =excel_max_row+1, column =1 ).value = book_title
            sh1.cell(row =excel_max_row+1, column =2 ).value = UPC
            sh1.cell(row =excel_max_row+1, column =3 ).value = product_type
            sh1.cell(row =excel_max_row+1, column =4 ).value = price_excltax
            sh1.cell(row =excel_max_row+1, column =5 ).value = price_incltax
            sh1.cell(row =excel_max_row+1, column =6 ).value = tax
            sh1.cell(row =excel_max_row+1, column =7 ).value = availability
            sh1.cell(row =excel_max_row+1, column =8 ).value = number_of_reviews
            wb.save(path)
            driver.back()
        
driver.quit()

